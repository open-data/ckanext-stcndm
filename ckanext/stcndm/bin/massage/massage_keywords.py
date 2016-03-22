#!/usr/bin/env python
# encoding: utf-8
import sys
from openpyxl import load_workbook
import json
from docopt import docopt


USAGE = """massage_keywords.py

Usage:
    insert_controlled_keywords.py <controlled.xlsx> <uncontrolled.xlsx>

"""
lang = {
    0L: u'en',
    1L: u'fr'
}


def main():
    args = docopt(USAGE)

    wb1 = load_workbook(args['<controlled.xlsx>'])
    keywords = {}
    for r in wb1.worksheets[0]:
        if not isinstance(r[0].value, long):
            continue
        keywords[r[0].value] = {
            u'title': {
                u'en': r[1].value,
                u'fr': r[2].value
            },
            u'keywords': {
                u'en': [r[1].value],
                u'fr': [r[2].value]
            }
        }
    wb2 = load_workbook(args['<uncontrolled.xlsx>'])
    for r in wb2.worksheets[0]:
        if not isinstance(r[0].value, long):
            continue
        keywords[r[0].value][u'keywords'][lang[r[2].value]].append(r[1].value)

    for keyword_id in keywords:
        print json.dumps({
            u'owner_org': u'statcan',
            u'private': False,
            u'type': u'keyword',
            u'name': u'keyword-'+unicode(keyword_id),
            u'title': keywords[keyword_id][u'title'],
            u'controlled_keyword_code': keyword_id,
            u'keywords': keywords[keyword_id][u'keywords']
        })


if __name__ == '__main__':
    sys.exit(main())
