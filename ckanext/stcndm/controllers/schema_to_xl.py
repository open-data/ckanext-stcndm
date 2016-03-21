# coding=UTF-8
from ckan.lib.base import response
from ckan.controllers.package import PackageController
from ckanext.scheming.logic import (scheming_dataset_schema_list,
                                    scheming_dataset_schema_show)
from openpyxl import Workbook, load_workbook
from openpyxl.cell import get_column_letter
from cStringIO import StringIO


class SchemaToXlController(PackageController):
    """
    Controller for sending schema in Excel file
    """

    def dump(self):
        concordance = {}
        wb = load_workbook('concordance.xlsx')
        for ws in wb.worksheets:
            dataset_type = ws.title
            for row in ws:
                refactored_search_name = row[1].value
                pre_refactored_search_name = row[0].value
                if dataset_type not in concordance:
                    concordance[dataset_type] = {}
                concordance[dataset_type][refactored_search_name] = pre_refactored_search_name
        schema_list = sorted(scheming_dataset_schema_list({}, {}))
        workbook = Workbook()
        sheet = None
        headers = [
            'ckan_field_name',
            'multivalued',
            'fluent',
            'code',
            'ckan_label_en',
            'ckan_label_fr',
            'search_name_en',
            'search_name_fr',
            'old_name',
            'old_name_en',
            'old_name_fr'
        ]
        for schema_name in schema_list:
            schema = []
            schema_dict = scheming_dataset_schema_show({}, {
                'type': schema_name,
                'expanded': True
            })
            concordance_dict = concordance.get(schema_name, {})
            for field in schema_dict['dataset_fields']:
                field_name = field.get('field_name')
                field_type = field.get('schema_field_type', '')
                fluent = field_type in [u'code', u'fluent']
                if fluent:
                    if field_type == u'code':
                        search_name_en = field_name + u'_desc_en'
                        search_name_fr = field_name + u'_desc_fr'
                    else:
                        search_name_en = field_name + u'_en'
                        search_name_fr = field_name + u'_fr'
                else:
                    search_name_en = u''
                    search_name_fr = u''

                schema.append(
                    {
                        'ckan_field_name': field_name,
                        'ckan_label_en': field.get(
                            'label',
                            {}
                        ).get('en', u'No label'),
                        'ckan_label_fr': field.get(
                            'label',
                            {}
                        ).get('fr', u'Pas d\'étiquette'),
                        'multivalued': field.get('schema_multivalued', False),
                        'fluent': fluent,
                        'code': field_type == u'code',
                        'search_name_en': search_name_en,
                        'search_name_fr': search_name_fr,
                        'old_name': concordance_dict.get(field_name, u' '),
                        'old_name_en': concordance_dict.get(search_name_en, u' '),
                        'old_name_fr': concordance_dict.get(search_name_fr, u' '),
                    }
                )
            if sheet:
                sheet = workbook.create_sheet(title=schema_name)
            else:
                sheet = workbook.active
                sheet.title = schema_name

            for i in range(1, len(headers)+1):
                cd = sheet.column_dimensions[get_column_letter(i)]
                sheet.cell(row=1, column=i, value=headers[i-1])
                cd.width = max(cd.width, len(headers[i-1]))
            row = 2
            for field in schema:
                column = 1
                for header in headers:
                    if isinstance(field[header], bool):
                        value = 'true' if field[header] else 'false'
                    else:
                        value = field[header]
                    cd = sheet.column_dimensions[get_column_letter(column)]
                    sheet.cell(row=row, column=column, value=value)
                    if value:
                        cd.width = max(cd.width, len(value))
                    column += 1
                row += 1
            cell = sheet['B2']
            sheet.freeze_panes = cell

        blob = StringIO()
        workbook.save(blob)
        response.headers['Content-Type'] = \
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = (
            'inline; filename="ProductSchemas.xlsx"')
        return blob.getvalue()
